import openpyxl
import json
import datetime

filename = "./CRA.xlsx"
entries = []

# This file takes a file called CRA.xlsx, with the specific worksheet called "for_encoding"
def read_from_file(filename, row_index):
    wb = openpyxl.load_workbook(filename)
    ws = wb["for_encoding"] 

    for row in range(2, row_index+1):
        entry = {}
        for col in range(1, ws.max_column+1):
            obj = ws.cell(row = row, column = col).value
            entry[ws.cell(row = 1, column = col).value] = obj
        entries.append(entry)

    #print(entries)

def date_serializer(obj):

    if isinstance(obj, datetime.datetime):
        return obj.isoformat()
    else:
        return obj


def export_file(filename):
    with open(filename, 'w') as fp:
        json.dump(entries, fp, indent=4, default=date_serializer)

read_from_file(filename,30)
export_file("output.json")
